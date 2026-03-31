"""Service helpers for storing fallback phone contacts when process/upload fails."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from threading import Lock

from openpyxl import Workbook, load_workbook

PHONE_PATTERN = r"^0(3|5|7|8|9)\d{8}$"
HEADERS = [
    "timestamp",
    "session_id",
    "phone",
    "error_reason",
    "frame_template_id",
    "selected_photos",
]

_write_lock = Lock()


class PhoneValidationError(ValueError):
    """Raised when a phone number is not a valid Vietnamese mobile number."""


class FallbackContactError(RuntimeError):
    """Raised when saving fallback contact information fails."""


@dataclass(slots=True)
class SavedFallbackContact:
    """Result metadata after successfully appending a fallback contact row."""

    file_path: Path
    row_index: int


def normalize_phone(phone: str) -> str:
    """Normalize phone input and validate the kiosk phone format."""
    normalized = "".join(ch for ch in phone if ch.isdigit())

    if not re.fullmatch(PHONE_PATTERN, normalized):
        raise PhoneValidationError(
            "Số điện thoại không hợp lệ. Vui lòng nhập số di động Việt Nam dạng 0xxxxxxxxx."
        )
    return normalized


def _get_daily_workbook_path(output_dir: Path, now: datetime) -> Path:
    filename = now.strftime("%Y-%m-%d.xlsx")
    return output_dir / filename


def append_fallback_contact(
    *,
    output_dir: Path,
    session_id: str,
    phone: str,
    error_reason: str,
    frame_template_id: str,
    selected_photos: list[str],
) -> SavedFallbackContact:
    """Append one fallback contact record into the daily workbook."""
    safe_phone = normalize_phone(phone)
    output_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    workbook_path = _get_daily_workbook_path(output_dir, now)

    try:
        with _write_lock:
            if workbook_path.exists():
                workbook = load_workbook(workbook_path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "fallback_contacts"
                sheet.append(HEADERS)

            selected_photos_text = ",".join(selected_photos)
            sheet.append(
                [
                    now.isoformat(timespec="seconds"),
                    session_id,
                    safe_phone,
                    error_reason,
                    frame_template_id,
                    selected_photos_text,
                ]
            )
            row_index = sheet.max_row
            workbook.save(workbook_path)

        return SavedFallbackContact(file_path=workbook_path, row_index=row_index)
    except PhoneValidationError:
        raise
    except Exception as exc:  # pragma: no cover - safeguard for IO layer
        raise FallbackContactError("Không thể lưu thông tin liên hệ dự phòng.") from exc
